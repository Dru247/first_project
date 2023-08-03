from openpyxl import load_workbook

from django.core.management.base import BaseCommand

from general_app.models import OperatorsSim, SimCards


class Command(BaseCommand):
    help = 'Проверка Симкарт СИМ2М по колличеству, между Проектои и сайтом' \
           ', с выводом недостающих симкарт'

    def handle(self, *args, **options):
        operator = OperatorsSim.objects.get(name='СИМ2М')
        sim_cards_project = SimCards.objects.filter(operator=operator)

        name_column = "ICC"
        name_file = "general_app/management/commands/export.xlsx"
        wb = load_workbook(filename = name_file)
        sheet = wb.active

        max_col = sheet.max_column
        number_column = 0
        exel_iccs = []

        for number_search_column in range(1, max_col + 1):
            cell_obj = sheet.cell(row = 1, column = number_search_column)
            if cell_obj.value == name_column:
                number_column = number_search_column

        for column_cell in sheet.iter_rows(
                min_row=2,
                min_col=number_column,
                max_col=number_column,
                values_only=True
                ):
            exel_iccs.append(column_cell[0])

        for sim in sim_cards_project:
            if sim.icc not in exel_iccs:
                print(f'{sim.icc} - на проекте есть, на сайте SIM2M нет')

        for sim_icc in exel_iccs:
            if SimCards.objects.filter(icc=sim_icc).exists():
                pass
            else:
                print(f'{sim_icc} - на SIM2M есть, на проекте нет')
